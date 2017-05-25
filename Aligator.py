#! /usr/bin/env python

#Automated Ligator = Aligator
#Version 1.0 (released 5/22/2017 by Patrick Erickson)

#Please visit the Aligator Github repository to view the legal license before using
#Aligator: https://github.com/kay-lab/Aligator

#This script takes all FASTA .txt files within a directory and predicts optimal ligation
#strategies. The top 1000 strategies (based on the scoring functions shown below) are
#shown within the Aligator analysis Excel file for each protein. All strategies are placed
#within text files in case the user wants to view them. All files listing all segments
#used to calculate the ligation strategies, along with their solubility scores, for
#the entire folder of FASTA .txt files is also shown within a separate text file.

#An intro to the user.
print "Welcome to Aligator!"
print ""

#Import important modules.
import platform
import re
import openpyxl #Needs to be installed by the user; creates formatted Excel files!
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from joblib import Parallel, delayed #Needs to be installed by the user; for parallel processing!
import multiprocessing
import glob
import os
import sys
import time
import shutil
from itertools import chain
import resource

#Imports os name (needed to determine safe mode method).
platName = platform.system()
platName = platName.lower()

#Creates the necessary output folders (If folders exist, tells users to delete them and try
#again).
try:
	os.makedirs("./Total Ligation Strategies Text Files")
except OSError:
	print "ERROR! A folder entitled 'Total Ligation Strategies Text Files' is already in"
	print "the current working directory. Please delete this folder or move the FASTA"
	print ".txt files to a different folder and try again!"
	print "Aligator terminated."
	print ""
	sys.exit()
try:
	os.makedirs("./All Segments with Solubility Scores")
except OSError:
	print "ERROR! A folder entitled 'All Segments with Solubility Scores' is already in"
	print "the current working directory. Please delete this folder or move the FASTA"
	print ".txt files to a different folder and try again!"
	print "Aligator terminated."
	print ""
	sys.exit()

#Creates colors to fill in Excel cells (openpyxl).
aquaFill = PatternFill(start_color='007FFFD4',
                       end_color='007FFFD4',
                       fill_type='solid')

greenFill = PatternFill(start_color='FF00FF00',
                       end_color='FF00FF00',
                       fill_type='solid')
                       
redFill = PatternFill(start_color='FFFF0000',
                       end_color='FFFF0000',
                       fill_type='solid')

#Setting for centering a cell in Excel (openpyxl).
center = Alignment(horizontal="center")

#The following codes for variables that are important in scoring segments and compiling
#optimal strategies. All of these variables cannot be changed while running Aligator,
#meaning that there are no user input prompts for the following variables. If you would
#like to view/change more of the scoring parameters, please see the "predict" function 
#written below.

#Lists of residues to be calculated for solubility. All based on our experiences with
#soluble amino acids and which residues are most problematic in preparing soluble
#peptides.
#List of positively-charged residues (good for solubility).
PosResList = ["K", "R", "H"]
#List of problematic-residues for solubility.
ProblematicResList = ["D", "E", "V", "I", "L"]

#The following solubility scores were found by observing the average solubility
#scores (solubility score / length of protein) for all segments for proteins involved
#in the E. coli 30S ribosomal subunit, 50S ribosomal subunit, and accessory/translation
#factors needed for the E. coli ribosome.
#Defines the expected average solubility score (point at which the score = 0).
meanSolLimit = -0.1581

#Defines the score that is one standard deviation from the expected average solubility 
#score (point at which the score = -1).
oneStdDev = -0.3128

#Defines the score that is two standard deviations from the expected average solubility 
#score (point at which the score = -2).
twoStdDev = -0.4675

#Defines the score that is three standard deviations from the expected average solubility 
#score (point at which the score = -3). All scores below this will be -3, as well.
threeStdDev = -0.6222

#Defines minimum length of segments allowed to use in finding ligation strategies.
MinSegLen = 10

#Defines the maximum number of strategies to put in the Excel output file:
MaxStrategies = 1000

#Defines the optimal segment length for scoring segments based on length (point at which
#score = 2)
bestSegmentLen = 40

#Defines the segment length used to cutoff ligation strategies that are no more than
#(protein length / CutoffSegLength) ligations long.
autoCutoffSegLength = 35

#Defines the segment length used to impose penalties for long ligation strategies, based on
#strategies that are longer than (protein length / penaltySegLength) segments.
autoPenaltySegLength = 40

#The following loop prompts allow the user to customize Aligator while running the script.
#Aligator will repeat the entered information to the user before performing synthesis
#strategy predictions, allowing the user to change entries if they were entered 
#incorrectly.
userInputInfo = False
while userInputInfo == False:
	#The following allows users to characterize thioesters differently than Aligator's default.
	#These statements explain the default thioester characterizations to the user.
	print "Aligator's default thioester characterization is mainly based on Fmoc hydrazide"
	print "SPPS compatibility, thereby forbidding segments with D, E, N, P, or Q"
	print "thioesters. The default scoring function for the acceptable segments is based"
	print "primarily on published NCL thioester kinetic rates."
	print ""
	
	print "Here are the default characterizations of thioesters:"
	#List of preferred thioesters (based on fastest thioester NCL kinetics).
	print "Preferred thioesters (score of +2): A, C, F, G, H, M, R, S, W, Y"
	#List of accepted thioesters (based on slower thioester NCL kinetics and K lactamization).
	print "Accepted thioesters (score of 0): I, K, L, T, V"
	#Forbidden Thioesters: D and E can undergo thioester migration to the side chain.
	#					   P thioesters have extremely slow kinetics.
	#					   D, N, and Q cannot be prepared by the hydrazide method.
	#Ligation sites with these thioesters will not be selected for scoring.
	print "Forbidden thioesters (segments CANNOT contain these): D, E, N, P, Q"
	print ""
	
	#Prompts the user for input regarding their thioester characterization choice.
	print "Would you like to keep the default thioester settings? If not, modify the"
	print "'Custom Thioester Input' Excel file to your choosing, and place this file into"
	print "the folder containing your FASTA text files."
	print ""
	
	customTEAns = raw_input("Enter 'yes' to keep the default, or enter 'no' to customize: ")
	print ""
	
	#Keeps default thioesters, if the user wishes to do so (keeps default if nothing entered).
	if customTEAns == "" or customTEAns[0].lower() == "y":
		PreferredTEList = ["A", "C", "F", "G", "H", "M", "R", "S", "W", "Y"]
		AcceptedTEList = ["I", "K", "L", "T", "V"] 
		ForbidTEList = ["D", "E", "N", "P", "Q"]
	#Everything in this 'else' statement attempts to load the thioester characterizations in
	#the input file, and if problems are detected, the user is told about the problem and
	#given a chance to change the input file appropriately.
	else:
		customTEEntryCheck = False
		while customTEEntryCheck == False:
			#Checks to make sure the input file is in the working directory.
			fileExistCheck = False
			while fileExistCheck == False:
				try:
					customTEFile = load_workbook(filename = 'Custom Thioester Input.xlsx',
					data_only=True)
					fileExistCheck = True
				except IOError:
					print "The 'Custom Thioester Input.xlsx' file is not in the current working"
					print "directory! Please put this file into the folder containing your FASTA"
					print "files and try again!"
					print ""
					checkpoint = raw_input("Press 'enter' when the input file is in the folder:")
					print ""
			
			#Puts the user's custom thioester characterizations into variables.
			sheet = customTEFile.get_sheet_by_name('Sheet1')
			customPref = str(sheet['B2'].value)
			customAccept = str(sheet['B3'].value)
			customForbid = str(sheet['B4'].value)
		
			#Removes any accidental white space in the cells of the workbook and makes letters
			#all uppercase. Also leaves blank if no thioesters are in the cell.
			customPref = customPref.upper().replace(" ", "").replace("NONE","")
			customAccept = customAccept.upper().replace(" ","").replace("NONE","")
			customForbid = customForbid.upper().replace(" ","").replace("NONE","")
			
			#The following variables all check the entries for mistakes.
			singleLetterCheck = True #Variable to check that single letter codes have only been entered.
			letterCheck = True #Variable to check incorrect thioester entries
			lengthCheck = True #Variable to check that all 20 thioester sites are characterized.
			
			#Checks to make sure only single letter AA abbreviations are entered.
			initialQualTECheck = customPref + "," + customAccept + "," + customForbid
			for i in initialQualTECheck.split(","):
				if len(i) > 1: #Could be 0, if no thioesters are in that category.
					print "ERROR! " + i + " is not formatted correctly. Please make sure"
					print "to seperate each single letter code with a comma in the input"
					print "file."
					print ""
					singleLetterCheck = False
			
			#Puts entries into one variable to allow for further quality checking of inputs.
			qualTECheck = (customPref.replace(",","") + customAccept.replace(",","")
						   + customForbid.replace(",",""))
			
			#This list enables checking that all canonical AA's have been categorized.
			AAList = ["A","C","D","E","F","G","H","I","K","L","M","N","P","Q","R","S",
					  "T","V","W","Y"]
			
			#Checks that all 20 canonical AAs have been classified as only one type of thioester.
			for i in qualTECheck:			
				if i in AAList:
					AAList.remove(i)
				else:
					print "ERROR! " + i + " has been entered more than once, or it is not"
					print "a canonical amino acid. Please fix this error in the input file"
					print "and try again."
					print ""
					letterCheck = False
			if len(AAList) != 0:
				print "ERROR! Not all thioesters have been classified!"
				for i in AAList:
					print i + " has not been characterized."
				print "Please fix this error in the input file and try again."
				print ""
				lengthCheck = False
			
			#Continues the user input options if no errors in the custom thioester input
			#file have been detected, or allows the user to fix them before continuing.
			if singleLetterCheck == True and letterCheck == True and lengthCheck == True:
				customTEEntryCheck = True
			else:
				checkpoint = raw_input("Press 'enter' when you have corrected and saved the input file:")
				print ""
		
		#Puts the thioesters into the appropriate lists needed for the rest of the program.
		PreferredTEList = customPref.split(",")
		AcceptedTEList = customAccept.split(",")
		ForbidTEList = customForbid.split(",")
		
		print "'Custom Thioester Input' file successfully read!"
		print ""
		
	#The following allows the user to change the maximum length allowed for segments.
	#Tells the user what the maximum length variable is used for and how to enter it.
	print "Please enter the maximum length (in number of residues) of peptide segments that" 
	print "can be considered in making ligation strategy predictions."
	print ""
	
	#Defines maximum length of segments allowed to use in finding ligation strategies. Also
	#checks to make sure that the user entered a proper response.
	validMaxSegLen = False
	while validMaxSegLen == False:
		try:
			MaxSegLen = int(raw_input("Enter the maximum segment length (only use numbers): "))
			print ""
			if MaxSegLen > MinSegLen:
				validMaxSegLen = True
			else:
				print "ERROR! The maximum segment length must be larger than the default"
				print "minimum segment length of " + str(MinSegLen) + " residues."
				print ""
		except ValueError:
			print "ERROR! That is not a valid entry. Please enter only numbers!"
			print ""

	#The following asks the user if they would like to turn on restriction mode, which 
	#cuts the number of possible segments to below 200, as well as extends both segment 
	#length cutoffs by 15 amino acids. These conditions were determined to work best through 
	#several tests of Aligator on different proteins within the E. coli ribosome.
	print "For large proteins, Aligator can initialize the recommended restriction mode,"
	print "which cuts the number of possible segments that can be used to create a"
	print "strategy to no more than 200. In addition, both segment lengths used to"
	print "determine the cutoff for number of ligations and imposing penalties on long"
	print "strategies will be increased by 15 amino acids, respectively, for proteins that"
	print "have more than 150 segments and are longer than 400 amino acids. This helps"
	print "reduce the computational costs of determining a ligation strategy, but may"
	print "result in no ligation strategies being found."
	print ""

	print "If this is your first time running a particular protein through Aligator,"
	print "we suggest you leave the default restriction mode on. If Aligator does not"
	print "find a ligation strategy, then re-run the program and turn off restriction"
	print "mode. The program will then enter 'safe mode,' which imposes a file size limit"
	print "to make sure that the script does not generate a huge amount of memory. If your"
	print "machine runs Mac OS X, then Aligator will stop running after 100 MB of memory"
	print "have been used by each processor. If your machine has another operating system,"
	print "then Aligator will stop running after 10 minutes in safe mode."
	print ""
	
	#Prompts the user to enter their decision (restriction mode turned on if nothing entered).
	print "Would you like to turn on restriction mode?"
	segmentReducerAns = raw_input("Enter 'yes' to turn on restriction mode. Enter 'no' to leave the mode off: ")
	print ""
	if segmentReducerAns == "" or segmentReducerAns[0].lower() == "y":
		segmentReducer = True
	else:
		segmentReducer = False

	#The following asks the user if they would like to have the helping hand solubility
	#reward implemented as part of the solubility scoring function.
	print "Helping hands can be installed onto Lys within peptide segments to dramatically"
	print "increase solubility. This script has an optional helping hand reward function,"
	print "which rewards segments containing a Lys residue by dividing the solubility"
	print "penalty by 2."
	print ""
	
	#Prompts the user to enter their decision (HH reward left on if nothing is entered).
	print "Would you like to include the helping hand reward function?"
	HHFlagAns = raw_input("Enter 'yes' to turn on the helping hand reward. Enter 'no' to leave this off: ")
	print ""
	if HHFlagAns == "" or HHFlagAns[0].lower() == "y":
		HHFlag = True
	else:
		HHFlag = False
	
	#Shows the user what they have chosen and allows them to loop back through the inputs
	#to change any mistakes.
	print "Here are the inputs that you have entered:"
	print ""
	
	#Generates variable of thioester lists to enable printing in one line.
	PrefOut = ' '.join(PreferredTEList)
	AcceptOut = ' '.join(AcceptedTEList)
	ForbidOut = ' '.join(ForbidTEList)
	
	#Prints thioester characterizations.
	print "THIOESTER CHARACTERIZATIONS"
	print "Preferred = " + PrefOut
	print "Accepted = " + AcceptOut
	print "Forbidden = " + ForbidOut
	print ""
	
	#Prints maximum segment length.
	print "MAXIMUM SEGMENT LENGTH ALLOWED"
	print str(MaxSegLen) + " residues"
	print ""
	
	#Prints the status of the restriction mode.
	print "RESTRICTION MODE STATUS"
	if segmentReducer == True:
		print "On"
	else:
		print "Off"
	print ""
	
	#Prints the status of the HH reward option.
	print "HELPING HAND REWARD STATUS"
	if HHFlag == True:
		print "On"
	else:
		print "Off"
	print ""
	
	#Allows the user to go back and change their inputs, if desired (goes on if nothing entered).
	continueAns = raw_input("Enter 'yes' to continue with these parameters. Enter 'no' to re-enter them: ")
	if continueAns == "" or continueAns[0].lower() == 'y':
		userInputInfo = True
	else:
		userInputInfo = False

#The following codes for functions important in making Aligator run properly.
#This allows for the FASTA .txt files to be entered into Aligator via ascending order.
numbers = re.compile(r'(\d+)')
def numericalSort(value):
	"""Function for sorting files in numerical and alphabetical ascending order"""
	parts = numbers.split(value)
	parts[1::2] = map(int, parts[1::2])
	return parts

#This is for combining the possible segments after scoring and adding up the scores.
def loop_rec(loopNum, segList, segScore, NtermPos, seg, finalCtermPos, cutoffSegLength, 
			 penaltySegLength, outList):
	"""Function for finding all possible ligation strategies and adding up score totals"""
	#Puts file size limit on the output by restricting how large outList can get, if
	#the user decides to enter safe mode.
	if segmentReducer == False:
		#This makes sure that the recursion doesn't account for ligations that are too long
		if loopNum <= (finalCtermPos+1)/cutoffSegLength:
			loopNum += 1
			for x in segList:
				#Limits loop by memory if Mac OS X detected
				if "darwin" in platName:
					if resource.getrusage(resource.RUSAGE_SELF).ru_maxrss <= 100  * 1024 * 1024: #100 MB
						#This saves strategies that have appended the final protein segment.
						if x[1] == NtermPos+1 and x[2] == finalCtermPos:
							newScore = segScore[0] + x[0][0]
							newTEScore = segScore[1] + x[0][1]
							newSolHHScore = segScore[2] + x[0][2]
							newLenScore = segScore[3] + x[0][3]
							newLigSiteScore = segScore[4] + x[0][4]
							newSeg = seg + "\t" + x[-1]
		
							#Penalty for an unideal amount of segments
							if loopNum > (finalCtermPos+1)/penaltySegLength:
								ligPenalty = -2*(loopNum-(((finalCtermPos+1)//penaltySegLength)))
								newScore += ligPenalty
							else:
								ligPenalty = 0
							
							#Puts ligation strategy into outList.
							newElement = [newScore, newTEScore, newSolHHScore, round(newLenScore,1), 
										  newLigSiteScore, ligPenalty, newSeg]
							outList.append(newElement)
							
							#Breaks the loop, killing that ligation search branch.
							break
						
						#This saves segments that are next to each other and then searches
						#for a segment that aligns with the newly-appended segment.
						elif x[1] == NtermPos+1 and x[2] != finalCtermPos:
							newScore = segScore[0] + x[0][0]
							newTEScore = segScore[1] + x[0][1]
							newSolHHScore = segScore[2] + x[0][2]
							newLenScore = segScore[3] + x[0][3]
							newLigSiteScore = segScore[4] + x[0][4]
							newTotalScore = [newScore, newTEScore, newSolHHScore, 
											 newLenScore, newLigSiteScore]
							newSeg = seg + "\t" + x[-1]
		
							loop_rec(loopNum, segList, newTotalScore, x[2], newSeg, 
									 finalCtermPos, cutoffSegLength, penaltySegLength, outList)
					#This breaks out of the strategy search loop once memory usage is too high.
					else:
						break
				
				#Limits loop by time if another os is detected.
				else:
					if time.time() - start_time <= 600: #10 minutes
						#This saves strategies that have appended the final protein segment.
						if x[1] == NtermPos+1 and x[2] == finalCtermPos:
							newScore = segScore[0] + x[0][0]
							newTEScore = segScore[1] + x[0][1]
							newSolHHScore = segScore[2] + x[0][2]
							newLenScore = segScore[3] + x[0][3]
							newLigSiteScore = segScore[4] + x[0][4]
							newSeg = seg + "\t" + x[-1]
	
							#Penalty for an unideal amount of segments
							if loopNum > (finalCtermPos+1)/penaltySegLength:
								ligPenalty = -2*(loopNum-(((finalCtermPos+1)//penaltySegLength)))
								newScore += ligPenalty
							else:
								ligPenalty = 0
							
							#Puts ligation strategy into outList.
							newElement = [newScore, newTEScore, newSolHHScore, round(newLenScore,1), 
										  newLigSiteScore, ligPenalty, newSeg]
							outList.append(newElement)
							
							#Breaks the loop, killing that ligation search branch.
							break
	
						#This saves segments that are next to each other and then searches
						#for a segment that aligns with the newly-appended segment.
						elif x[1] == NtermPos+1 and x[2] != finalCtermPos:
							newScore = segScore[0] + x[0][0]
							newTEScore = segScore[1] + x[0][1]
							newSolHHScore = segScore[2] + x[0][2]
							newLenScore = segScore[3] + x[0][3]
							newLigSiteScore = segScore[4] + x[0][4]
							newTotalScore = [newScore, newTEScore, newSolHHScore, 
											 newLenScore, newLigSiteScore]
							newSeg = seg + "\t" + x[-1]
	
							loop_rec(loopNum, segList, newTotalScore, x[2], newSeg, 
									 finalCtermPos, cutoffSegLength, penaltySegLength, outList)
					
					#This breaks the ligation strategy search loop once the time limit is up.
					else:
						break
	
	#This is what occurs when restriction mode is left on by the user.
	else:
		#This makes sure that the recursion doesn't account for ligations that are too long
		if loopNum <= (finalCtermPos+1)/cutoffSegLength:
			loopNum += 1
			for x in segList:
				#This saves strategies that have appended the final protein segment.
				if x[1] == NtermPos+1 and x[2] == finalCtermPos:
					newScore = segScore[0] + x[0][0]
					newTEScore = segScore[1] + x[0][1]
					newSolHHScore = segScore[2] + x[0][2]
					newLenScore = segScore[3] + x[0][3]
					newLigSiteScore = segScore[4] + x[0][4]
					newSeg = seg + "\t" + x[-1]
				
					#Penalty for an unideal amount of segments
					if loopNum > (finalCtermPos+1)/penaltySegLength:
						ligPenalty = -2*(loopNum-(((finalCtermPos+1)//penaltySegLength)))
						newScore += ligPenalty
					else:
						ligPenalty = 0
					
					#Puts ligation strategy into outList.
					newElement = [newScore, newTEScore, newSolHHScore, round(newLenScore,1), 
								  newLigSiteScore, ligPenalty, newSeg]
					outList.append(newElement)
					
					#Breaks the loop, killing that ligation search branch.
					break
				
				#This saves segments that are next to each other and then searches
				#for a segment that aligns with the newly-appended segment.
				elif x[1] == NtermPos+1 and x[2] != finalCtermPos:
					newScore = segScore[0] + x[0][0]
					newTEScore = segScore[1] + x[0][1]
					newSolHHScore = segScore[2] + x[0][2]
					newLenScore = segScore[3] + x[0][3]
					newLigSiteScore = segScore[4] + x[0][4]
					newTotalScore = [newScore, newTEScore, newSolHHScore, newLenScore, 
									 newLigSiteScore]
					newSeg = seg + "\t" + x[-1]
				
					loop_rec(loopNum, segList, newTotalScore, x[2], newSeg, finalCtermPos, 
							 cutoffSegLength, penaltySegLength, outList)
	return outList #returns all ligation strategies found to outList variable (see below)

#This is the main function of the program that predicts optimal ligation strategies based
#on just the FASTA text file that is given.
def predict(File, AverageSolubScoreList):
	"""Function for predicting optimal total chemical ligation strategies for proteins"""
	inFile = open(File, "U") #U makes sure there are no unicode errors in .txt files.
	ProteinName = File.rstrip(".txt") #Saves the name of the protein.
	if ProteinName.endswith("fasta"): #Cuts off 'fasta' from the file name, if there.
		ProteinName = ProteinName.rstrip(".fasta")
		
	#Tells user which protein is being analyzed.
	print "Now analyzing " + ProteinName
	print ""
	
	#Creates separate sheet for the protein being analyzed.
	outFile.create_sheet(index=-1, title=ProteinName)
	sheet = outFile.get_sheet_by_name(ProteinName)
	
	#Skips the info line in the fasta sequence.
	inFile.readline()
	
	#Places the entire sequence into the variable ProteinSeq.
	ProteinSeq = inFile.read()
	ProteinSeq = ProteinSeq.replace("\n","").replace("\r","").replace(" ","").replace("\t","")
	ProteinSeq = ProteinSeq.upper()
	ProteinLength = len(ProteinSeq) #Important for segment indexing.
	
	#Checks to see if more than 1 fasta sequence is in the .txt file.
	#If more than 1 is found, the program quits and gives the user an error message.
	if ">" in ProteinSeq:
		print "ERROR! More than 1 fasta sequence was detected in your "+ProteinName+" file!"
		print "Aligator does not take .txt files with more than 1 fasta sequence."
		print "Please reformat this file and try again."
		print "Aligator terminated!"
		print ""
		sys.exit()
	
	#Finds indexes of Cys or Ala and saves them in a list (needed for saving thioester
	#indexes in next part of script).
	LigIndexList = ([pos for pos, char in enumerate(ProteinSeq) 
					if char == "C" or char == "A"])
	
	#Find indexes of thioesters and saves them in a list.
	TEIndexList = []
	for i in LigIndexList:
		if i != 0:
			TEIndexList.append(i-1)
	
	#Splits amino acid sequence up via Cys and Ala sites.
	ProteinRead = re.split("([C]|[A])", ProteinSeq) #This code keeps C and A in list.
	if ProteinRead[0] != "":
		FirstSeg = [ProteinRead[0]] #Need to save this segment or it gets lost!
	else:
		FirstSeg = []
	#The following puts C and A back into their correct segments.
	ProteinRead = [i+j for i,j in zip(ProteinRead[1::2], ProteinRead[2::2])]
	#Puts the first segment back into the ProteinRead segment list.
	ProteinRead = FirstSeg + ProteinRead
	
	#Creates initial segment array.
	InitSegments = []
	ProteinReadCounter = 0
	NtermIndex = 0
	for i in TEIndexList:
		InitSegments.append([NtermIndex,i,ProteinRead[ProteinReadCounter]])
		ProteinReadCounter += 1
		NtermIndex = i + 1
	#This saves the last segment (which doesn't have a thioester) in the array.
	InitSegments.append([NtermIndex,ProteinLength-1, ProteinRead[-1]])
	
	#Combines segments that have forbidden thioesters, removing segments with these
	#thioesters from being considered in the calculations.
	ForbidTECounter = 0
	NoForbidTESegments = []
	for i in InitSegments[:-1]:
		if i[2][-1] in ForbidTEList:
			InitSegments[ForbidTECounter+1][2] = i[2] + InitSegments[ForbidTECounter+1][2]
			InitSegments[ForbidTECounter+1][0] = i[0]
		else:
			NoForbidTESegments.append(i)
		ForbidTECounter += 1
	NoForbidTESegments.append(InitSegments[-1]) #Saves final segment in new array.
	
	#Creates the final segment list by combining all possible segments that are
	#less than or equal to the maximum segment length.
	CombinedSegments = []
	finalIndexNum = len(NoForbidTESegments)
	iterationTracker = 1 #Allows all segments in NoForbidTESegments to go through loop.
	for i in NoForbidTESegments:
		CombinedSegments.append(i)
		NtermIndex = i[0]
		SegCounter = iterationTracker - 1 #tracks segs next to seg currently being analyzed.
		currentIndexNum = iterationTracker
		if currentIndexNum < finalIndexNum:
			SegCounter += 1
			CtermIndex = NoForbidTESegments[SegCounter][1]
			PotentNewSeg = (i[2]) + NoForbidTESegments[SegCounter][2]
			while len(PotentNewSeg) <= MaxSegLen and currentIndexNum < finalIndexNum:
				CombinedSegments.append([NtermIndex, CtermIndex, PotentNewSeg])
				currentIndexNum += 1
				SegCounter += 1
				if currentIndexNum < finalIndexNum:
					PotentNewSeg = PotentNewSeg + NoForbidTESegments[SegCounter][2]
					CtermIndex = NoForbidTESegments[SegCounter][1]
		iterationTracker += 1
	
	#Removes any segments that are smaller than the MinSegLen variable.
	FinalSegments = []
	for i in CombinedSegments:
		if len(i[-1]) >= MinSegLen:
			FinalSegments.append(i)
	
	#Decreases the number of segments being scored and counted in the possible
	#ligation strategy calculations by sequentially trimming from the smallest and largest
	#segments allowed until no more than 200 segments exist. This only works if the user
	#leaves the restriction mode on!
	if len(FinalSegments) > 200 and segmentReducer == True:
		newMinSegLen = MinSegLen
		newMaxSegLen = MaxSegLen
		while len(FinalSegments) > 200:
			newMinSegLen += 1
			for i in FinalSegments:
				if len(i[-1]) < newMinSegLen:
					FinalSegments.remove(i)
			if len(FinalSegments) > 200:
				newMaxSegLen -= 1
				for i in FinalSegments:
					if len(i[-1]) > newMaxSegLen:
						FinalSegments.remove(i)
		
		print "To reduce the amount of memory usage, the minimum segment length cutoff was"
		print "increased from "+str(MinSegLen)+ " amino acids to "+str(newMinSegLen)
		print "amino acids. The maximum segment length cutoff was reduced from "+str(MaxSegLen)
		print "amino acids to "+str(newMaxSegLen)+" amino acids."
		print ""
		print "This reduced the number of possible segments to "+str(len(FinalSegments))+"."
		print ""
		
	#Increases the segment lengths used to create the ligation number cutoff and penalty
	#for strategies containing too many segments. This only happens if the user leaves the
	#restriction mode on, as well as if the protein is over 400 amino acids long and
	#has over 150 possible segments!
	if len(FinalSegments) > 150 and ProteinLength > 400 and segmentReducer == True:
		cutoffSegLength = autoCutoffSegLength + 15
		penaltySegLength = autoPenaltySegLength + 15
		
		print "Since there are "+str(len(FinalSegments))+" possible segments and the protein"
		print "is over 400 amino acids long, both segment lengths will be increased by 15."
		print "As a result, the segment length used to make the cutoff for number of"
		print "ligations in a strategy is now "+str(cutoffSegLength)+" while the segment length"
		print "used to penalize strategies greater than a certain number of segments is now "
		print str(penaltySegLength)+"."
		
	else:
		cutoffSegLength = autoCutoffSegLength
		penaltySegLength = autoPenaltySegLength
	
	#The following are the scoring functions for all viable segments.
	ScoredFragments = []
	for i in FinalSegments:
		Score = 0
		
		#Score based on thioesters in segments.
		TEScore = 0
		if i[1] != ProteinLength-1: #This causes the C-terminal protein segments to not be counted.
			if i[-1][-1] in PreferredTEList:
				Score += 2
				TEScore += 2
			elif i[-1][-1] in AcceptedTEList:
				Score += 0
				TEScore += 0
		
		#Creates an average solubility score for each segment.
		SolubScore = 0
		HHSite = False
		segLen = len(i[-1])
		for char in i[-1]:
			if char in PosResList:
				SolubScore += 1
				if char == "K":
					HHSite = True
			elif char in ProblematicResList:
				SolubScore += (-1)
		#Saves each segment with their average solubility score to create an output
		#text file later.
		AverageSolubScore = (float(SolubScore) / segLen)
		AverageSolubScoreList.append([i[-1], AverageSolubScore])
		
		#Scale for scoring segments based on average solubility.
		#Scores based on average solubility distributions observed for 3 different
		#protein subsets of the E. coli ribosome.
		if AverageSolubScore >= meanSolLimit:
			FinalSolubScore = 0
		elif AverageSolubScore < meanSolLimit and AverageSolubScore >= (oneStdDev):
			FinalSolubScore = (-1)*((AverageSolubScore - (meanSolLimit)) / ((oneStdDev)-(meanSolLimit)))
		elif AverageSolubScore < (oneStdDev) and AverageSolubScore >= (twoStdDev):
			FinalSolubScore = (-1) + ((-1)*((AverageSolubScore - (oneStdDev)) / ((twoStdDev)-(oneStdDev))))
		elif AverageSolubScore < (twoStdDev) and AverageSolubScore >= (threeStdDev):
			FinalSolubScore = (-2) + ((-1)*((AverageSolubScore - (twoStdDev)) / ((threeStdDev)-(twoStdDev))))
		elif AverageSolubScore < (threeStdDev):
			FinalSolubScore = (-3)
			
		#The following rewards the presence of a Lys for potentially insoluble
		#fragments (due to availability of a helping hand on Lys), if the user wants.
		SolubleHHScore = 0
		if FinalSolubScore < 0 and HHSite == True and HHFlag == True:
			Score += (float(FinalSolubScore) / 2)
			SolubleHHScore += (float(FinalSolubScore) / 2)
		else:
			Score += FinalSolubScore
			SolubleHHScore += FinalSolubScore
				
		#Score based on length of segment.
		lenScore = 0
		if len(i[-1]) == bestSegmentLen:
			Score += 2
			lenScore += 2
		elif len(i[-1]) < bestSegmentLen:
			Score += 2 + ((len(i[-1]) - bestSegmentLen) * 0.1)
			lenScore += 2 + ((len(i[-1]) - bestSegmentLen) * 0.1)
		else:
			Score += 2 + ((len(i[-1]) - bestSegmentLen) * -0.1)
			lenScore += 2 + ((len(i[-1]) - bestSegmentLen) * -0.1)
		
		#Score based on whether segment has Cys or Ala ligation site.
		LigSiteScore = 0
		if i[0] != 0: #This will cause the beginning N-terminal segments to not be counted
			if i[-1].startswith("A"):
				Score += (-2)
				LigSiteScore += (-2)
		
		Entry = [[Score] + [TEScore] + [SolubleHHScore] + [lenScore] + [LigSiteScore]] + i
		ScoredFragments.append(Entry)
	
	#Adds up scores and creates ligation strategies using the recursive loop function.
	print "Now finding all possible ligation strategies"
	print ""
	
	#Stores scores and sequences for whole segments into bigList, then removes them from
	#scored segment list (important with small proteins).
	wholeSegList = []
	for x in ScoredFragments:
		if x[1] == 0 and x[2] == ProteinLength-1:
			wholeSegList.append([x[0][0],x[0][1],x[0][2],x[0][3],x[0][4], 0, x[-1]])
			ScoredFragments.remove(x)
	
	#Allows parallel processing, in case safe mode is turned on.
	num_cores = multiprocessing.cpu_count()
	
	#Calls recursive loop function.
	#This does not use parallel processing, as parallel processing causes some memory
	#issues for larger proteins in restriction mode!
	if segmentReducer == True:
		bigList = (loop_rec(1,ScoredFragments,i[0],i[2],i[-1],ProteinLength-1,cutoffSegLength,
					penaltySegLength,[]) for i in ScoredFragments if i[1] == 0)
	#This calls parallel processing, as it makes scanning through the different strategies
	#more efficient when there is a data limit set when restriction mode is turned off.
	else:
		bigList = Parallel(n_jobs=num_cores)(delayed(loop_rec)(1,ScoredFragments,i[0],
					  i[2],i[-1],ProteinLength-1,cutoffSegLength,penaltySegLength,[])
					  for i in ScoredFragments if i[1] == 0)
	
	#Sorts bigList by score.
	print "Now sorting"
	print ""
	
	#Needed to remove a dimension from the array generated from the recursive function.
	bigList = list(chain.from_iterable(bigList))
	
	#Adds whole segments back into the possible ligation strategies.
	bigList = bigList + wholeSegList
	
	#Actually sorts bigList.
	bigList.sort(key=lambda x: float(x[0]), reverse=True)
	
	#Writes output files.
	print "Writing output files"
	print ""
	
	#Puts starting info into the total ligation strategy .txt file.
	txtFile = open(ProteinName+" All Strategies.txt", "w") #Stores all strategies.
	txtFile.write("Strategy Score\tSegments\n")
	
	stratTracker = 0 #Keeps track of how many strategies have been written into Excel.
	longestStrat = 0 #Keeps track of longest ligation strategy.
	rowNum = 2
	columnNum = 1
	#If no ligation strategies were found, this writes that into the output files.
	if len(bigList) == 0:
		bigList.append(["NA","NA","NA","NA","NA","NA","NO LIGATION STRATEGIES FOUND!"])
	for i in bigList:
		#Needed to only write max number of top strategies into Excel file
		if stratTracker < MaxStrategies:
			#Writes info into the text file.
			txtFile.write(str(i[0])+"\t"+i[-1]+"\n")
			
			#Writes scores for a strategy into the Excel file.
			cell = sheet.cell(row = rowNum, column = columnNum)
			cell.value = i[0]
			columnNum += 1
			cell = sheet.cell(row = rowNum, column = columnNum)
			cell.value = i[1]
			columnNum += 1
			cell = sheet.cell(row = rowNum, column = columnNum)
			cell.value = i[2]
			columnNum += 1
			cell = sheet.cell(row = rowNum, column = columnNum)
			cell.value = i[3]
			columnNum += 1
			cell = sheet.cell(row = rowNum, column = columnNum)
			cell.value = i[4]
			columnNum += 1
			cell = sheet.cell(row = rowNum, column = columnNum)
			cell.value = i[5]
			columnNum += 1
			
			#Writes each segment within a strategy into the Excel file.
			for segment in i[-1].split("\t"):
				if longestStrat < columnNum: #Helps format Excel sheet later
					longestStrat = columnNum
				cell = sheet.cell(row = rowNum, column = columnNum)
				cell.value = segment
				columnNum += 1
			stratTracker += 1
			rowNum += 1
			columnNum = 1
		
		#Once the max number of top strategies have been written, only writes strategies
		#to the text file.
		else:
			txtFile.write(str(i[0])+"\t"+i[-1]+"\n")
	
	#Writes initial information into Excel file and formats the file.
	sheet["A1"] = "Total Strategy Score"
	sheet["A1"].alignment = center
	sheet["A1"].font = Font(size=12, bold=True)
	sheet["A1"].fill = greenFill
	
	sheet["B1"] = "Thioester Total Score"
	sheet["B1"].alignment = center
	sheet["B1"].font = Font(size=12, bold=True)
	sheet["B1"].fill = redFill
	
	sheet["C1"] = "Solubility Total Score"
	sheet["C1"].alignment = center
	sheet["C1"].font = Font(size=12, bold=True)
	sheet["C1"].fill = redFill
	
	sheet["D1"] = "Segment Length Total Score"
	sheet["D1"].alignment = center
	sheet["D1"].font = Font(size=12, bold=True)
	sheet["D1"].fill = redFill
	
	sheet["E1"] = "Total Ala Junction Site Penalty"
	sheet["E1"].alignment = center
	sheet["E1"].font = Font(size=12, bold=True)
	sheet["E1"].fill = redFill
	
	sheet["F1"] = "Total Penalty for # of Ligations"
	sheet["F1"].alignment = center
	sheet["F1"].font = Font(size=12, bold=True)
	sheet["F1"].fill = redFill
	
	sheet.merge_cells(start_row=1, start_column=7, end_row=1, end_column=longestStrat)
	sheet["G1"] = "Segments (from N- to C-terminus)"
	sheet["G1"].alignment = center
	sheet["G1"].font = Font(size=12, bold=True)
	sheet["G1"].fill = aquaFill
	
	#Closes total strategy text file.
	txtFile.close()
	
	#Moves the total strategy text file into the correct folder.
	shutil.move("./" + ProteinName + " All Strategies.txt", 
				"./Total Ligation Strategies Text Files/" + ProteinName + " All Strategies.txt")
			
	#Close the input FASTA text file.
	inFile.close()
	
	return AverageSolubScoreList #Allows segment text file to be written.

#The rest of the script actually executes everything!
#Lets user know that segment predictions have started.
print "Aligator will now predict ideal synthesis strategies!"
print ""

#Records starting time.
start_time = time.time()

#Creates the output Excel file that will have the ligation strategies for each FASTA file.
outFile = openpyxl.Workbook()

#Loop through each FASTA text file to predict ligation strategies.
EmptyList = []
for File in sorted(glob.iglob("*.txt"), key=numericalSort):
	AvgSolubScoreList = predict(File, EmptyList)

#Removes a blank sheet in the output Excel file.
outFile.remove_sheet(outFile.get_sheet_by_name("Sheet"))

#Writes average solubility score list to a text file.
AvgSoluScoreFile = open("All Segments with Solubility Scores.txt", "w")
#Checks to make sure FASTA text files were in the directory.
try:
	for i in AvgSolubScoreList:
		AvgSoluScoreFile.write(str(i[0])+"\t"+str(i[1])+"\n")
	AvgSoluScoreFile.close()
except NameError:
	print "It appears that no FASTA files are in your current directory!"
	print "Please add these files to the directory and try again."
	print "Aligator terminated."
	sys.exit()

#Moves the average solubility score file to a folder.
shutil.move("./All Segments with Solubility Scores.txt", 
			"./All Segments with Solubility Scores/All Segments with Solubility Scores.txt")

#Saves output Excel file.
cwd = os.getcwd()
folder = os.path.basename(cwd)
#Tries to save Excel file and gives user an error message if there were no input files
#in the starting directory.
try:
	outFile.save("Aligator Analysis for " + folder + ".xlsx")
except IndexError:
	print "ERROR: It appears that there are no FASTA text files in the directory!"
	print "Please put FASTA text files into the directory and try again."
	print "Aligator terminated!"
	sys.exit()

#Prints conclusion to user and lists full time it took to run Aligator.
print "Aligator complete!"
print ""
print "Aligator took %s seconds to run." % round((time.time() - start_time), 2)