#! /usr/bin/env python3

#Automated Ligator = Aligator

#Aligator: https://github.com/kay-lab/Aligator

#Version 2.0 (GitHub Release Date TBD)

#This script takes all FASTA .txt files within a directory and predicts optimal ligation
#strategies. The top 1000 strategies (based on the scoring functions shown below) are
#shown within the Aligator analysis Excel file for each protein. All strategies are placed
#within text files in case the user wants to view them. All viable peptide segments
#used to calculate the ligation strategies, along with their solubility scores, for
#each protein are shown within the Viable Segment Lists Excel file. The Aligator Run
#Information document shows all user inputs, statistics, and run time for the
#program. All output files are stored within a folder that is timestamped based on when
#Aligator was launched.

#IMPORTANT NOTE: In order to make Aligator function properly as an executable file, line
#51 was added. Please ensure that this line is disabled if you are not using the
#executable.

#An intro to the user.
print ("")
print ("Welcome to Aligator!")
#Date and time
import datetime
now = datetime.datetime.now()
print ("Current date and time: ")
print (now.strftime("%Y-%m-%d %H:%M:%S"))
print ("")

#Import important modules.
import re
import csv
import openpyxl #Needs to be installed by the user; creates formatted Excel files!
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import glob
import os
import sys
import time
import shutil


#The following changes the working directory to the folder in which the Python executable
#is stored in. Disable this if you are not using the executable.
# os.chdir(os.path.dirname(sys.executable))

#Gets the name of the current working directory (for file naming later)
cwd = os.getcwd()
folder = os.path.basename(cwd) #Needed to get name of working directory.

#Creates the output folder for a run based on the timestamp.
timestamp = str(datetime.datetime.now().strftime("%B %d, %Y %I_%M_%S %p"))
os.makedirs("./" + timestamp)
OutputFolder=f'./{timestamp}'


# The following variables toggle different functions of the program (for development purposes only).

output_all_strategies_text=True # Gives "_ All Strategies.txt" file output in a sub-folder.

unrestrained_mode=False # Removes the dead-end elimination method in strategy-building. Provides mathematically equivalent results at higher processing cost. Leave False unless wishing to compare dead-end elimination vs. original method.

report_to_screen=True # Detailed progress is reported to screen during Aligator processing loop.

prompt_for_user_inputs=True # Toggle False to quickly test program without running user interface, using default values for all inputs.

merge_output_csv=True # Final output CSV files are merged into Excel format (.xlsx) for better readability.


#The following codes for variables that are important in scoring segments and compiling
#optimal strategies. All of these variables cannot be changed while running Aligator,
#meaning that there are no user input prompts for the following variables.

#Lists of residues to be calculated for solubility. All based on our experiences with
#soluble amino acids and which residues are most problematic in preparing soluble
#peptides.
#List of positively-charged residues (good for solubility).
PosResList = ["K", "R", "H"]
#List of problematic-residues for solubility.
ProblematicResList = ["D", "E", "V", "I", "L"]

#The following solubility scores were found by observing the average solubility
#scores (solubility score / length of protein) for all segments for proteins involved
#in the E. coli 30S ribosomal subunit, 50S ribosomal subunit, and accessory/translation
#factors needed for the E. coli ribosome.
#Defines the expected average solubility score (point at which the score = 0).
meanSolLimit = -0.1581

#Defines the score that is one standard deviation from the expected average solubility
#score (point at which the score = -1).
oneStdDev = -0.3128

#Defines the score that is two standard deviations from the expected average solubility
#score (point at which the score = -2).
twoStdDev = -0.4675

#Defines the score that is three standard deviations from the expected average solubility
#score (point at which the score = -3). All scores below this will be -3, as well.
threeStdDev = -0.6222

#Defines minimum length of segments allowed to use in finding ligation strategies.
MinSegLen = 10

#Defines the maximum number of strategies to put in the Excel output file:
MaxStrategies = 1000

#Defines the optimal segment length for scoring segments based on length (point at which
#score = 2)
bestSegmentLen = 40

#Defines the segment length used to cutoff ligation strategies that are no more than
#(protein length / CutoffSegLength) ligations long.
autoCutoffSegLength = 35

#Defines the segment length used to impose penalties for long ligation strategies, based on
#strategies that are longer than (protein length / penaltySegLength) segments.
autoPenaltySegLength = 40

#Creates colors to fill in Excel cells (openpyxl).
aquaFill = PatternFill(start_color='007FFFD4',
                       end_color='007FFFD4',
                       fill_type='solid')

greenFill = PatternFill(start_color='FF00FF00',
                       end_color='FF00FF00',
                       fill_type='solid')

redFill = PatternFill(start_color='FFFF0000',
                       end_color='FFFF0000',
                       fill_type='solid')

#Setting for centering a cell in Excel (openpyxl).
center = Alignment(horizontal="center")


# FUNCTION DEFS
#The following codes for functions important in making Aligator run properly.


# This function is useful for printing time of operations to screen
def gettime(reporttext):
    print(f'{reporttext} -  {time.time()-start_time} seconds')

#This allows for the FASTA .txt files to be entered into Aligator via ascending order.
numbers = re.compile(r'(\d+)')
def numericalSort(value):
    """Function for sorting files in numerical and alphabetical ascending order"""
    parts = numbers.split(value)
    parts[1::2] = map(int, parts[1::2])
    return parts

# This function takes an input strategy (a list of numbers representing start, NCL junctions, and end) and gives a dictionary of scores
def scoreStrategy(InputStrategy):
    ScoreDict={
        "thioester":0,
        "solubility":0,
        "length":0,
        "thiol":0,
        "ligations":0,
        "total":0
        }
    # Loop through segments and add up their sub-scores
    for i in range(0,len(InputStrategy)-1):
        SegmentLeft=InputStrategy[i]
        SegmentRight=InputStrategy[i+1]
        SegmentKey=(SegmentLeft,SegmentRight)
        ScoreDict["thioester"]+=SegmentScoreDict[SegmentKey]["thioester"]
        ScoreDict["solubility"]+=SegmentScoreDict[SegmentKey]["solubility"]
        ScoreDict["length"]+=SegmentScoreDict[SegmentKey]["length"]
        ScoreDict["thiol"]+=SegmentScoreDict[SegmentKey]["thiol"]
    # Calculate 'ideal' number of segments in a strategy; strategies with more segments than this will be penalized
    # As an example, if this ends up being 6.5 (e.g., a 260-aa protein with ideal segment length of 40), 6-segment and below strategies will be fine, 7+ will be penalized
    # Save as an integer (cutting off the decimal and rounding down) for easier math
    ProteinLength=InputStrategy[-1]-InputStrategy[0]
    IdealSegmentCount = int(ProteinLength/autoPenaltySegLength)
    # Add ligation penalty if necessary
    NumberOfSegments=len(InputStrategy)-1
    if NumberOfSegments>IdealSegmentCount:
        LigationPenalty = -2 * (NumberOfSegments-IdealSegmentCount)
        ScoreDict["ligations"]+=LigationPenalty
    # Sum up total score
    SumScore=ScoreDict["thioester"]+ScoreDict["solubility"]+ScoreDict["length"]+ScoreDict["thiol"]+ScoreDict["ligations"]
    ScoreDict["total"]=SumScore
    # Return dictionary of scores
    return ScoreDict


# CREATE RUN INFO FILE
# Info about the run will periodically be written to this file; do not open or edit this file
# while Aligator is actively running.
# To avoid issues with file permissions, if user prompts are enabled then this file is instead
# opened after all inputs have been accepted.
if prompt_for_user_inputs==False:
    RunInfoFile = open(f"{OutputFolder}/Aligator Run Information.doc", "w")

#USER INPUT PROMPTS
#The following loop prompts allow the user to customize Aligator while running the script.
#Aligator will repeat the entered information to the user before performing synthesis
#strategy predictions, allowing the user to change entries if they were entered
#incorrectly.
userInputInfo = False
while userInputInfo == False and prompt_for_user_inputs == True:
    #The following allows users to characterize thioesters differently than Aligator's default.
    #These statements explain the default thioester characterizations to the user.
    print ("Aligator's default thioester characterization is mainly based on Fmoc hydrazide")
    print ("SPPS compatibility, thereby forbidding segments with D, E, N, P, or Q")
    print ("thioesters. The default scoring function for the acceptable segments is based")
    print ("primarily on published NCL thioester kinetic rates.")
    print ("")

    print ("Here are the default characterizations of thioesters:")
    #List of preferred thioesters (based on fastest thioester NCL kinetics).
    print ("Preferred thioesters (score of +2): A, C, F, G, H, M, R, S, W, Y")
    #List of accepted thioesters (based on slower thioester NCL kinetics and K lactamization).
    print ("Accepted thioesters (score of 0): I, K, L, T, V")
    #Forbidden Thioesters: D and E can undergo thioester migration to the side chain.
    #                      P thioesters have extremely slow kinetics.
    #                      D, N, and Q cannot be prepared by the hydrazide method.
    #Ligation sites with these thioesters will not be selected for scoring.
    print ("Forbidden thioesters (segments CANNOT contain these): D, E, N, P, Q")
    print ("")

    #Prompts the user for input regarding their thioester characterization choice.
    print ("Would you like to keep the default thioester settings? If not, modify the")
    print ("'Custom Parameters Input' Excel file to your choosing, and place this file into")
    print ("the folder containing your FASTA text files.")
    print ("")

    customTEAns = input("Enter 'yes' to keep the default, or enter 'no' to customize: ")
    print ("")

    #Keeps default thioesters, if the user wishes to do so (keeps default if nothing entered).
    if customTEAns == "" or customTEAns[0].lower() == "y":
        PreferredTEList = ["A", "C", "F", "G", "H", "M", "R", "S", "W", "Y"]
        AcceptedTEList = ["I", "K", "L", "T", "V"]
        ForbidTEList = ["D", "E", "N", "P", "Q"]
    #Everything in this 'else' statement attempts to load the thioester characterizations in
    #the input file, and if problems are detected, the user is told about the problem and
    #given a chance to change the input file appropriately.
    else:
        customTEEntryCheck = False
        while customTEEntryCheck == False:
            #Checks to make sure the input file is in the working directory.
            fileExistCheck = False
            while fileExistCheck == False:
                try:
                    customParametersFile = load_workbook(filename = 'Custom Parameters Input.xlsx',
                    data_only=True)
                    fileExistCheck = True
                except IOError:
                    print ("The 'Custom Parameters Input.xlsx' file is not in the current working")
                    print ("directory! Please put this file into the folder containing your FASTA")
                    print ("files and try again!")
                    print ("")
                    checkpoint = input("Press 'enter' when the input file is in the folder:")
                    print ("")

            #Puts the user's custom thioester characterizations into variables.
            sheet = customParametersFile["Sheet1"]
            customPref = str(sheet['B2'].value)
            customAccept = str(sheet['B3'].value)
            customForbid = str(sheet['B4'].value)

            #Removes any accidental white space in the cells of the workbook and makes letters
            #all uppercase. Also leaves blank if no thioesters are in the cell.
            customPref = customPref.upper().replace(" ", "").replace("NONE","")
            customAccept = customAccept.upper().replace(" ","").replace("NONE","")
            customForbid = customForbid.upper().replace(" ","").replace("NONE","")

            #The following variables all check the entries for mistakes.
            singleLetterCheck = True #Variable to check that single letter codes have only been entered.
            letterCheck = True #Variable to check incorrect thioester entries
            lengthCheck = True #Variable to check that all 20 thioester sites are characterized.

            #Checks to make sure only single letter AA abbreviations are entered.
            initialQualTECheck = customPref + "," + customAccept + "," + customForbid
            for i in initialQualTECheck.split(","):
                if len(i) > 1: #Could be 0, if no thioesters are in that category.
                    print ("ERROR! " + i + " is not formatted correctly. Please make sure")
                    print ("to separate each single letter code with a comma in the input")
                    print ("file.")
                    print ("")
                    singleLetterCheck = False

            #Puts entries into one variable to allow for further quality checking of inputs.
            qualTECheck = (customPref.replace(",","") + customAccept.replace(",","")
                           + customForbid.replace(",",""))

            #This list enables checking that all canonical AA's have been categorized.
            AAList = ["A","C","D","E","F","G","H","I","K","L","M","N","P","Q","R","S",
                      "T","V","W","Y"]

            #Checks that all 20 canonical AAs have been classified as only one type of thioester.
            for i in qualTECheck:
                if i in AAList:
                    AAList.remove(i)
                else:
                    print ("ERROR! " + i + " has been entered more than once, or it is not")
                    print ("a canonical amino acid. Please fix this error in the input file")
                    print ("and try again.")
                    print ("")
                    letterCheck = False
            if len(AAList) != 0:
                print ("ERROR! Not all thioesters have been classified!")
                for i in AAList:
                    print (i + " has not been characterized.")
                print ("Please fix this error in the input file and try again.")
                print ("")
                lengthCheck = False

            #Continues the user input options if no errors in the Custom Parameters Input
            #file have been detected, or allows the user to fix them before continuing.
            if singleLetterCheck == True and letterCheck == True and lengthCheck == True:
                customTEEntryCheck = True
            else:
                checkpoint = input("Press 'enter' when you have corrected and saved the input file:")
                print ("")

        #Puts the thioesters into the appropriate lists needed for the rest of the program.
        PreferredTEList = customPref.split(",")
        AcceptedTEList = customAccept.split(",")
        ForbidTEList = customForbid.split(",")

        print ("Custom thioesters - file successfully read!")
        print(f'Preferred: {", ".join(PreferredTEList)}')
        print(f'Accepted: {", ".join(AcceptedTEList)}')
        print(f'Forbidden: {", ".join(ForbidTEList)}')
        print ("")


    #Similar to above, but for thiol sites
    #The following allows users to characterize thiol sites to an expanded list beyond C or A
    #These statements explain the default thiol characterizations to the user.
    print ("By default, Aligator assumes that segments are joined by NCL between a")
    print ("thioester and thiol. Typically Cys is used as the thiol, but desulfurization")
    print ("enables Ala and other possible sites; however, most thiolated AAs other than")
    print ("C or A have poor ligation kinetics.")
    print ("")

    print ("Here are the default allowed thiols:")
    #List of preferred thiol sites, no penalty (typically only C).
    print ("Preferred thiols (no penalty): C")
    #List of accepted thiols, with only a desulfurization penalty (typically only A).
    print ("Accepted thiols (-2 desulfurization penalty): A")
    #List of poor thiols, with kinetics and desulfurization penalties (typically blank, but V
    #is relatively common in literature).
    print ("Poor thiols (-4 desulfurization & kinetics penalty): (None)")
    print ("")

    #Prompts the user for input regarding their thiol characterization choice.
    print ("Would you like to keep the default thiol sites? If not, modify the")
    print ("'Custom Parameters Input' Excel file to your choosing, and place this file into")
    print ("the folder containing your FASTA text files.")
    print ("")

    customThiolAns = input("Enter 'yes' to keep the default, or enter 'no' to customize: ")
    print ("")

    #Keeps default thioesters, if the user wishes to do so (keeps default if nothing entered).
    if customThiolAns == "" or customThiolAns[0].lower() == "y":
        GoodThiolList=['C']
        OKThiolList=['A']
        PoorThiolList=[]

    #Everything in this 'else' statement attempts to load the thioester characterizations in
    #the input file, and if problems are detected, the user is told about the problem and
    #given a chance to change the input file appropriately.
    else:
        customEntryCheck = False
        while customEntryCheck == False:
            #Checks to make sure the input file is in the working directory.
            fileExistCheck = False
            while fileExistCheck == False:
                try:
                    customParametersFile = load_workbook(filename = 'Custom Parameters Input.xlsx',
                    data_only=True)
                    fileExistCheck = True
                except IOError:
                    print ("The 'Custom Parameters Input.xlsx' file is not in the current working")
                    print ("directory! Please put this file into the folder containing your FASTA")
                    print ("files and try again!")
                    print ("")
                    checkpoint = input("Press 'enter' when the input file is in the folder:")
                    print ("")

            #Puts the user's custom thiol characterizations into variables.
            sheet = customParametersFile["Sheet1"]
            customGood = str(sheet['B6'].value)
            customOK = str(sheet['B7'].value)
            customPoor = str(sheet['B8'].value)

            #Removes any accidental white space in the cells of the workbook and makes letters
            #all uppercase. Also leaves blank if no thioesters are in the cell.
            customGood = customGood.upper().replace(" ", "").replace("NONE","")
            customOK = customOK.upper().replace(" ","").replace("NONE","")
            customPoor = customPoor.upper().replace(" ","").replace("NONE","")

            #The following variables all check the entries for mistakes.
            singleLetterCheck = True #Variable to check that single letter codes have only been entered.
            letterCheck = True #Variable to check incorrect thiol entries
            lengthCheck = True #Variable to check that AT LEAST ONE thiol is characterized

            #Checks to make sure only single letter AA abbreviations are entered.
            initialQualThiolCheck = customGood + "," + customOK + "," + customPoor
            for i in initialQualThiolCheck.split(","):
                if len(i) > 1: #Could be 0, if no thiols are in that category.
                    print ("ERROR! " + i + " is not formatted correctly. Please make sure")
                    print ("to separate each single letter code with a comma in the input")
                    print ("file.")
                    print ("")
                    singleLetterCheck = False

            #Puts entries into one variable to allow for further quality checking of inputs.
            qualThiolCheck = (customGood.replace(",","") + customOK.replace(",","")
                           + customPoor.replace(",",""))

            #This list enables checking that only canonical AA's have been categorized.
            AAList = ["A","C","D","E","F","G","H","I","K","L","M","N","P","Q","R","S",
                      "T","V","W","Y"]

            #Checks that the list is not blank
            if qualThiolCheck=="":
                print("ERROR! No thiols detected in any category.")
                print('Fill at least one "thiol" field in the input file')
                print("and try again.")
                print("")
                lengthCheck=False
            #Checks that only canonical AAs have been classified as only one type of thiol
            for i in qualThiolCheck:
                if i in AAList:
                    AAList.remove(i)
                else:
                    print ("ERROR! " + i + " has been entered more than once, or it is not")
                    print ("a canonical amino acid. Please fix this error in the input file")
                    print ("and try again.")
                    print ("")
                    letterCheck = False

            #Continues the user input options if no errors in the Custom Parameters Input
            #file have been detected, or allows the user to fix them before continuing.
            if singleLetterCheck == True and letterCheck == True and lengthCheck == True:
                customEntryCheck = True
            else:
                checkpoint = input("Press 'enter' when you have corrected and saved the input file:")
                print ("")

        #Puts the thioesters into the appropriate lists needed for the rest of the program.
        GoodThiolList = customGood.split(",")
        OKThiolList = customOK.split(",")
        PoorThiolList = customPoor.split(",")

        print ("Custom thiols - file successfully read!")
        print(f'Preferred: {", ".join(GoodThiolList)}')
        print(f'Accepted: {", ".join(OKThiolList)}')
        print(f'Poor: {", ".join(PoorThiolList)}')
        print ("")


    #The following asks the user if they would like to have the helping hand solubility
    #reward implemented as part of the solubility scoring function.
    print ("Helping hands can be installed onto Lys and Glu side chains within peptide")
    print ("segments to dramatically increase solubility. This script has an optional")
    print ("helping hand reward function, which rewards segments containing Lys or Glu by")
    print ("dividing the solubility penalty by 2.")
    print ("")

    #Prompts the user to enter their decision (HH reward left on if nothing is entered).
    print ("Would you like to include the helping hand reward function? If you decide to")
    print ("include the helping hand reward function, you will have the option to")
    print ("customize the solubility enhancement attachment sites.")
    print ("")
    HHFlagAns = input("Enter 'yes' to turn on the helping hand reward. Enter 'no' to leave this off: ")
    print ("")
    if HHFlagAns == "" or HHFlagAns[0].lower() == "y":
        HHFlag = True
    else:
        HHFlag = False

    #if the user turned the HH reward on:
    #The following prompts the user for input regarding their HH characterization choice.

    if HHFlag == True:

        print ("Would you like to keep the default Lys and Glu residues as helping hand")
        print ("attachment sites? If not, modify the 'Custom Parameters Input' Excel file")
        print ("to your choosing, and place this file into the folder containing your")
        print ("FASTA text files.")
        print ("")

        customHHSiteAns = input("Enter 'yes' to keep the default, or enter 'no' to customize: ")
        print ("")

        #Keeps default HH amino acids, if the user wishes to do so (keeps default if nothing entered).
        if customHHSiteAns == "" or customHHSiteAns[0].lower() == "y":
            SolubilizingTagList = ["K", "E"]


        #Everything in this 'else' statement attempts to load the HH site characterizations in
        #the input file, and if problems are detected, the user is told about the problem and
        #given a chance to change the input file appropriately.
        else:
            customHHSiteEntryCheck = False
            while customHHSiteEntryCheck == False:
                #Checks to make sure the input file is in the working directory.
                fileExistCheck = False
                while fileExistCheck == False:
                    try:
                        customParametersFile = load_workbook(filename = 'Custom Parameters Input.xlsx',
                        data_only=True)
                        fileExistCheck = True
                    except IOError:
                        print ("The 'Custom Parameters Input.xlsx' file is not in the current")
                        print ("working directory! Please put this file into the folder")
                        print ("containing your FASTA files and try again!")
                        print ("")
                        checkpoint = input("Press 'enter' when the input file is in the folder:")
                        print ("")

                #Puts the user's custom HH characterizations into variables.
                sheet = customParametersFile["Sheet1"]
                customHHSite = str(sheet['B11'].value)

                #Removes any accidental white space in the cells of the workbook and makes letters
                #all uppercase. Also leaves blank if no AAs are in the cell.
                customHHSite = customHHSite.upper().replace(" ", "").replace("NONE","")

                #The following variables all check the entries for mistakes.
                singleLetterCheck = True #Variable to check that single letter codes have only been entered.
                letterCheck = True #Variable to check incorrect HH entries

                #Checks to make sure only single letter AA abbreviations are entered.
                initialHHSiteCheck = customHHSite + ","
                for i in initialHHSiteCheck.split(","):
                    if len(i) > 1: #Could be 0, if no amino acids are in that category.
                        print ("ERROR! " + i + " is not formatted correctly. Please make sure")
                        print ("to separate each single letter code with a comma in the input")
                        print ("file.")
                        print ("")
                        singleLetterCheck = False

                #Puts entries into one variable to allow for further quality checking of inputs.
                qualHHCheck = customHHSite.replace(",","")

                #This list enables checking that only canonical AAs have been selected and that none are selected more than once.
                AAList = ["A","C","D","E","F","G","H","I","K","L","M","N","P","Q","R","S",
                          "T","V","W","Y"]

                #Checks that only canonical AAs have been selected.
                for i in qualHHCheck:
                    if i in AAList:
                        AAList.remove(i)
                    else:
                        print ("ERROR! " + i + " has been entered more than once, or it is not")
                        print ("a canonical amino acid. Please fix this error in the input file")
                        print ("and try again.")
                        print ("")
                        letterCheck = False

                #Continues the user input options if no errors in the custom parameter input
                #file have been detected, or allows the user to fix them before continuing.
                if singleLetterCheck == True and letterCheck == True:
                    customHHSiteEntryCheck = True
                else:
                    checkpoint = input("Press 'enter' when you have corrected and saved the input file:")
                    print ("")

            #Puts the HHs into the appropriate lists needed for the rest of the program.
                SolubilizingTagList = customHHSite.split(",")

            print ("'Custom Parameters Input' file successfully read!")
            print (f"Attachment Sites = {customHHSite}")
            print ("")

    #The following allows the user to change the maximum length allowed for segments.
    #Tells the user what the maximum length variable is used for and how to enter it.
    print ("Please enter the maximum length (in number of residues) of peptide segments that")
    print ("can be considered in making ligation strategy predictions.")
    print ("")

    #Defines maximum length of segments allowed to use in finding ligation strategies. Also
    #checks to make sure that the user entered a proper response.
    validMaxSegLen = False
    while validMaxSegLen == False:
        try:
            MaxSegLen = int(input("Enter the maximum segment length (only use numbers): "))
            print ("")
            if MaxSegLen > MinSegLen:
                validMaxSegLen = True
            else:
                print ("ERROR! The maximum segment length must be larger than the default")
                print ("minimum segment length of " + str(MinSegLen) + " residues.")
                print ("")
        except ValueError:
            print ("ERROR! That is not a valid entry. Please enter only numbers!")
            print ("")

    #Shows the user what they have chosen and allows them to loop back through the inputs
    #to change any mistakes.
    print ("Here are the inputs that you have entered:")
    print ("")

    #Generates variable of AA lists to enable printing in one line.
    PrefOut = ' '.join(PreferredTEList)
    AcceptOut = ' '.join(AcceptedTEList)
    ForbidOut = ' '.join(ForbidTEList)
    if HHFlag == True:
        PrefHHSite = ' '.join(SolubilizingTagList)
    GoodOut = ' '.join(GoodThiolList)
    OKOut = ' '.join(OKThiolList)
    PoorOut = ' '.join(PoorThiolList)

    #Prints thioester characterizations.
    print ("THIOESTER CHARACTERIZATIONS")
    print ("Preferred = " + PrefOut)
    print ("Accepted = " + AcceptOut)
    print ("Forbidden = " + ForbidOut)
    print ("")

    #Prints thiol characterizations.
    print ("THIOL SITES")
    print ("Preferred = " + GoodOut)
    print ("Accepted = "+OKOut)
    print ("Poor = "+PoorOut)
    print ("")

    #Prints the status of the HH reward option.
    print ("HELPING HAND REWARD STATUS")
    if HHFlag == True:
        print ("On")
        print ("")
        print ("HELPING HAND CHARACTERIZATIONS")
        print (PrefHHSite)

    else:
        print ("Off")
    print ("")

    #Prints maximum segment length.
    print ("MAXIMUM SEGMENT LENGTH ALLOWED")
    print (str(MaxSegLen) + " residues")
    print ("")


    #Allows the user to go back and change their inputs, if desired (goes on if nothing entered).
    continueAns = input("Enter 'yes' to continue with these parameters. Enter 'no' to re-enter them: ")
    if continueAns == "" or continueAns[0].lower() == 'y':
        userInputInfo = True

        #The following writes the input options to an Aligator run info file.
        cwd = os.getcwd()
        folder = os.path.basename(cwd) #Needed to get name of working directory.

        RunInfoFile = open(f"{OutputFolder}/Aligator Run Information.doc", "w")

        RunInfoFile.write("Aligator Run Information for "+folder+"\n")
        RunInfoFile.write("\n")
        RunInfoFile.write("Aligator Initiated on "+timestamp+"\n")
        RunInfoFile.write("\n")
        RunInfoFile.write("USER INPUTS:\n")
        RunInfoFile.write("THIOESTER CHARACTERIZATIONS:\n")
        RunInfoFile.write("Preferred = "+PrefOut+"\n")
        RunInfoFile.write("Accepted = "+AcceptOut+"\n")
        RunInfoFile.write("Forbidden = "+ForbidOut+"\n")
        RunInfoFile.write("\n")
        RunInfoFile.write("HELPING HAND REWARD STATUS:\n")
        if HHFlag == True:
            RunInfoFile.write("On\n")
            RunInfoFile.write("\n")
            RunInfoFile.write("HELPING HAND CHARACTERIZATIONS:\n")
            RunInfoFile.write(PrefHHSite+"\n")
            RunInfoFile.write("\n")
        else:
            RunInfoFile.write("Off\n")
        RunInfoFile.write("\n")
        RunInfoFile.write("MAXIMUM SEGMENT LENGTH ALLOWED:\n")
        RunInfoFile.write(str(MaxSegLen)+" residues\n")
        RunInfoFile.write("\n")

    else:
        userInputInfo = False
# Done gathering user input variables

# Developer mode only; if this toggle is False, the prompts above will be ignored
# and instead the following default values will be used:
if not prompt_for_user_inputs:
    PreferredTEList = ["A", "C", "F", "G", "H", "M", "R", "S", "W", "Y"]
    AcceptedTEList = ["I", "K", "L", "T", "V"]
    ForbidTEList = ["D", "E", "N", "P", "Q"]
    GoodThiolList=['C']
    OKThiolList=['A']
    PoorThiolList=['V']
    MaxSegLen=60
    HHFlag=True
    SolubilizingTagList=['K','E']
    # Write to Run Info file; it's not as nicely formatted, but that's what you get
    # for being a developer.
    RunInfoFile.write('NO USER INPUTS, DEFAULT VALUES USED\n')
    RunInfoFile.write(f'PreferredTEList = {PreferredTEList}\n')
    RunInfoFile.write(f'AcceptedTEList = {AcceptedTEList}\n')
    RunInfoFile.write(f'ForbidTEList = {ForbidTEList}\n')
    RunInfoFile.write(f'GoodThiolList = {GoodThiolList}\n')
    RunInfoFile.write(f'OKThiolList = {OKThiolList}\n')
    RunInfoFile.write(f'PoorThiolList = {PoorThiolList}\n')
    RunInfoFile.write(f'MaxSegLen = {MaxSegLen}\n')
    RunInfoFile.write(f'HHFlag = {HHFlag}\n')
    RunInfoFile.write(f'SolubilizingTagList = {SolubilizingTagList}\n')
    RunInfoFile.write(f'\n')

#The rest of the script actually executes everything!
#Lets user know that segment predictions have started.
print ("Aligator will now predict ideal synthesis strategies!")
print ("")

#Records starting time.
start_time = time.time()


#Loop through each FASTA text file to get a list of valid sequences (with protein names)
ProteinNameAndSeqList = []
for Filename in sorted(glob.iglob("*.txt"), key=numericalSort):
    # Get protein name
    ProteinName = Filename.rstrip(".txt") #Saves the name of the protein.
    if ProteinName.endswith("fasta"): #Cuts off 'fasta' from the file name, if there.
        ProteinName = ProteinName.rstrip(".fasta")
    with open(Filename, 'r') as inFile:
        #Reads the info line in the fasta sequence.
        FirstLine = inFile.readline()

        #Places the entire sequence into the variable ProteinSeq.
        ProteinSeq = inFile.read()
        ProteinSeq = ProteinSeq.replace("\n","").replace("\r","").replace(" ","").replace("\t","")
        ProteinSeq = ProteinSeq.upper()
        ProteinLength = len(ProteinSeq)

        #Checks to see if there is information written below the first line, as well as if
        #descriptor character (>) is in first line. Sometimes, fasta files downloaded from
        #websites would not be in proper fasta format (1st line should be a description and
        #the remaining lines are all protein sequence), so this check prevents this problem.
        if FirstLine[0] != ">" or ProteinSeq == "":
            print ("ERROR! The "+ProteinName+" file is not in proper fasta format! Please make")
            print ("sure that all protein fasta files have a single description line with the")
            print ("> symbol at the beginning, followed by separate lines containing only the")
            print ("amino acid sequence of the protein!")
            print ("Aligator terminated!")
            print ("")
            sys.exit()

        #Checks to see if more than 1 fasta sequence is in the .txt file.
        #If more than 1 is found, the program quits and gives the user an error message.
        if ">" in ProteinSeq:
            print ("ERROR! More than 1 fasta sequence was detected in your "+ProteinName+" file!")
            print ("Aligator does not take .txt files with more than 1 fasta sequence.")
            print ("Please reformat this file and try again.")
            print ("Aligator terminated!")
            print ("")
            sys.exit()

    # If all looks correct, save this protein name and sequence to our list
    ProteinNameAndSeqList.append((ProteinName,ProteinSeq))

# Quit if no fasta sequences are detected; otherwise list each one
if len(ProteinNameAndSeqList)==0:
    print(f'ERROR: No FASTA sequences detected in current folder "{folder}" !')
    print(f'Place at least one .txt or .fasta file in the current folder and try again.')
    print ("")
    sys.exit()
print(f'Detected {len(ProteinNameAndSeqList)} fasta sequences:')
for (ProteinName,ProteinSeq) in ProteinNameAndSeqList:
    print(ProteinName)
print("")


# Keep track of output csv files made for each protein
SegFileDict={}
LigFileDict={}
MaxWidthDict={} # Tracks the largest # segments in any output strategy, for Excel formatting

# Loop through each sequence to generate the required output files: Valid Segments (.csv), Aligator Analysis (.csv), and All Strategies (.txt)
# After all loops are complete, CSV files of the same type will be merged into a single Excel document and formatted
for (ProteinName,ProteinSeq) in ProteinNameAndSeqList:
    print(f'Now running {ProteinName} ({len(ProteinSeq)} aa)...')

    # Remember the longest strategy for Excel formatting later; by default this is 1 segment
    MaxWidthSoFar=1
    MaxWidthDict[ProteinName]=1

    # SEGMENT CALCULATIONS AND SCORING
    # Calculate 'ideal' number of segments in a strategy; strategies with more segments than this will be penalized
    # As an example, if this ends up being 6.5 (e.g., a 260-aa protein with ideal segment length of 40), 6-segment and below strategies will be fine, 7+ will be penalized
    # Save as an integer (cutting off the decimal and rounding down) for easier math later
    IdealSegmentCount = int(len(ProteinSeq)/autoPenaltySegLength)

    # Determine the index of all valid ligation junctions (i.e., not forbidden thioesters, and more than MinSegLen away from either end)
    # These will be used to generate all possible segments within the protein
    SegmentBorderList=[0] # Beginning of protein counts as a segment border
    ThiolList=GoodThiolList+OKThiolList+PoorThiolList
    for i,Char in enumerate(ProteinSeq):
        if Char in ThiolList and MinSegLen<=i<=(len(ProteinSeq)-MinSegLen) and not ProteinSeq[i-1] in ForbidTEList:
            SegmentBorderList.append(i)
    # Add a marker for the end of the protein as well
    SegmentBorderList.append(len(ProteinSeq))

    # Determine from this list if any strategies will be possible; if there is a long stretch in between ligation junctions with no valid segments, we cannot make strategies
    StrategiesArePossible=True
    LastPosition=0
    for n in SegmentBorderList:
        if LastPosition+MaxSegLen<n:
            StrategiesArePossible=False
            print(f'NO POSSIBLE STRATEGIES - Large gap between junctions {LastPosition} and {n} ({n-LastPosition}-aa segment)')
        if LastPosition+MinSegLen<n:
            LastPosition=n

    if report_to_screen==True:
        gettime('Scoring all possible segments')

    # Define all possible segments for the protein, discarding those too small to be considered. If not too large, score and add to dictionary.
    SegmentScoreDict={} # Only includes valid segments between minimum and maximum length
    StartPointDict={} # Contains all segments grouped by starting point
    for i,LeftIndex in enumerate(SegmentBorderList):
        for RightIndex in SegmentBorderList[i+1:]:
            Segment=ProteinSeq[LeftIndex:RightIndex]
            # If segment is an acceptable size, score and add to segment dictionary for later reference
            if MinSegLen<=len(Segment)<=MaxSegLen:
                SegmentKey=(LeftIndex,RightIndex)
                # This is a valid start point; keep track of all segments sharing this start point
                if not LeftIndex in StartPointDict:
                    StartPointDict.update({LeftIndex:[]})
                StartPointDict[LeftIndex].append(RightIndex) # List of right indices of valid segments
                SegmentScoreDict.update({SegmentKey:{}}) # Key is index borders of segment; value is a sub-dictionary detailing all scores
                SegmentScoreDict[SegmentKey].update({'seq':Segment}) # Add segment sequence

                #Score based on thioesters in segments.
                TEScore = 0
                if RightIndex != len(ProteinSeq): #This causes the C-terminal protein segments to not be counted.
                    if ProteinSeq[RightIndex-1] in PreferredTEList:
                        TEScore += 2
                    elif ProteinSeq[RightIndex-1] in AcceptedTEList:
                        TEScore += 0
                SegmentScoreDict[SegmentKey].update({'thioester':TEScore})

                #Creates an average solubility score for each segment.
                SolubScore = 0
                HHSite = False
                for char in Segment:
                    if char in PosResList:
                        SolubScore += 1
                    elif char in ProblematicResList:
                        SolubScore -= 1
                    if HHFlag == True and char in SolubilizingTagList:
                        HHSite = True
                SegmentScoreDict[SegmentKey].update({'HH':HHSite})
                #Divide by length to get average
                AverageSolubScore = (float(SolubScore) / len(Segment))
                #Scale for scoring segments based on average solubility.
                #Scores based on average solubility distributions observed for 3 different
                #protein subsets of the E. coli ribosome.
                if AverageSolubScore >= meanSolLimit:
                    FinalSolubScore = 0
                elif AverageSolubScore < meanSolLimit and AverageSolubScore >= (oneStdDev):
                    FinalSolubScore = (-1)*((AverageSolubScore - (meanSolLimit)) / ((oneStdDev)-(meanSolLimit)))
                elif AverageSolubScore < (oneStdDev) and AverageSolubScore >= (twoStdDev):
                    FinalSolubScore = (-1) + ((-1)*((AverageSolubScore - (oneStdDev)) / ((twoStdDev)-(oneStdDev))))
                elif AverageSolubScore < (twoStdDev) and AverageSolubScore >= (threeStdDev):
                    FinalSolubScore = (-2) + ((-1)*((AverageSolubScore - (twoStdDev)) / ((threeStdDev)-(twoStdDev))))
                elif AverageSolubScore < (threeStdDev):
                    FinalSolubScore = (-3)
                # If helping hand reward function is on, negative solubility scores are halved
                if FinalSolubScore < 0 and HHSite == True and HHFlag == True:
                    FinalSolubScore = float(FinalSolubScore)/2
                # Save solubility score to dictionary
                SegmentScoreDict[SegmentKey].update({'solubility':FinalSolubScore}) # This is actually used for strategy scoring
                SegmentScoreDict[SegmentKey].update({'avgsolubility':AverageSolubScore}) # Both are reported in Segment Scores output file

                #Score based on length of segment.
                lenScore = 0
                if len(Segment) == bestSegmentLen:
                    lenScore += 2
                elif len(Segment) < bestSegmentLen:
                    lenScore += 2 + ((len(Segment) - bestSegmentLen) * 0.1)
                else:
                    lenScore += 2 + ((len(Segment) - bestSegmentLen) * -0.1)
                # Save length score to dictionary
                SegmentScoreDict[SegmentKey].update({'length':lenScore})

                # Thiol penalty - apply penalty for desulfurization (e.g., Ala), and double penalty for poor kinetics w/desulfurization (e.g., Val)
                # May be changed in Custom Parameters Input file
                LigSiteScore = 0
                if LeftIndex!=0: # This causes the leftmost segment to not be counted
                    if Segment[0] in OKThiolList:
                        LigSiteScore-=2
                    elif Segment[0] in PoorThiolList:
                        LigSiteScore-=4
                SegmentScoreDict[SegmentKey].update({'thiol':LigSiteScore})

                # TOTAL SCORE - SUM OF ALL OTHER SCORES
                TotalScore=TEScore+FinalSolubScore+lenScore+LigSiteScore
                SegmentScoreDict[SegmentKey].update({'total':TotalScore})

    gettime(f'Segment scoring complete...found {len(SegmentScoreDict)} valid segments')

    # Write segment solubility scores to output file
    OutputFilepath=f'{OutputFolder}/Viable Segment List for {ProteinName}.csv'
    with open(OutputFilepath,'w') as f:
        # Write header line
        f.write(f'First AA,Last AA,Sequence,Average AA Solubility,Final Solubility Score,Solubility Tag Sites ({"/".join(SolubilizingTagList)})?\n')
        # If no viable segments, write n/a in relevant fields
        if len(SegmentScoreDict)==0:
            f.write('n/a,n/a,NO VIABLE SEGMENTS')
        # Otherwise loop through segments
        else:
            for (FirstAA,LastAA) in sorted(SegmentScoreDict.keys()):
                SegmentKey=(FirstAA,LastAA)
                # Get text to report for true/false value for helping hand sites
                HHReportText=""
                if SegmentScoreDict[SegmentKey]["HH"]==True:
                    HHReportText="Yes"
                # Write info about segment to its own line
                f.write(f'{str(FirstAA+1)},{str(LastAA)},{SegmentScoreDict[SegmentKey]["seq"]},{SegmentScoreDict[SegmentKey]["avgsolubility"]},{SegmentScoreDict[SegmentKey]["solubility"]},{HHReportText}\n')

    # Report success and track for later
    print('Wrote "Viable Segment List" file')
    SegFileDict[ProteinName]=OutputFilepath
    print('-----------------')

    # 'Unrestrained' mode - temporarily set the Max Strategies to a ridiculously high number; but continue to trim the Excel output file
    # Only recommended for development purposes
    if unrestrained_mode==True:
        OldMaxStrategies=MaxStrategies
        MaxStrategies=1000000000000000000000000000000000

    # Begin creating strategies, resulting in final sorted list which will be written to file
    if StrategiesArePossible==True:
        print('Now creating strategies....')

        # Create the starting list of segments to begin processing all possible strategies
        StrategyQueue = []
        FinalStrategyList = []
        for r in StartPointDict[0]:
            Strategy = (0,r) # Single-segment strategy looks the same as a single segment
            StrategyQueue.append(Strategy)

        if report_to_screen==True:
            print(f'Found {len(StrategyQueue)} starting segments')

        # For printing to screen
        loopcount=0

        # MAIN LOOP OF BUILDING STRATEGIES
        while len(StrategyQueue)>0:
            loopcount+=1
            if report_to_screen==True:
                gettime(f'-------\nRanking {loopcount+1}-segment strategies...\n')
            # Copy and clear queue
            PrevQueue=StrategyQueue[:]
            report_inputstrats=len(PrevQueue)
            StrategyQueue=[]
            NextQueue=[]
            # Loop through copied list to generate all strategies with 1 additional segment
            report_finalstrats=0
            for Strategy in PrevQueue:
                LastAA=Strategy[-1]
                # If this strategy is complete (ends at the final AA), add it to our final output list
                if LastAA==len(ProteinSeq):
                    FinalStrategyList.append(Strategy)
                    report_finalstrats=len(FinalStrategyList)
                # If the strategy is incomplete, generate the list of next strategies by adding 1 additional segment
                elif LastAA in StartPointDict:
                    NextEndPoints=StartPointDict[LastAA]
                    for EndPoint in NextEndPoints:
                        NextStrategy=list(Strategy)+[EndPoint] # Convert to list of numbers and add next number
                        NextStrategy=tuple(NextStrategy) # Convert back to tuple
                        NextQueue.append(NextStrategy)

            # Done adding strategies to queue; verbose printout for debugging
            # if report_to_screen==True:
                # print(f'Detected {report_finalstrats} complete strategies (total {len(FinalStrategyList)} so far), {report_inputstrats-report_finalstrats} partial')
                # print(f'Partial list expanded to {len(NextQueue)} next strategies')
            report_newstratsfound=len(NextQueue)


            # SORT AND TRIM PARTIAL STRATEGIES
            # Verbose printout for debugging
            # if report_to_screen==True:
            #         print(f'Performing dead-end elimination... (<={MaxStrategies} strategies per endpoint)')
            # Partial strategy scores can only be directly compared if they represent the same slice of the protein; in other words, if they share a start and endpoint
            # Since all partial strategies share the same start point (0), group these by endpoint
            PartialStrategiesByEndPoint={}
            for Strategy in NextQueue:
                EndPoint=Strategy[-1]
                if not EndPoint in PartialStrategiesByEndPoint:
                    PartialStrategiesByEndPoint.update({EndPoint:[]}) # First time we encounter a number, add a blank entry
                PartialStrategiesByEndPoint[EndPoint].append(Strategy) # Save each strategy with its endpoint

            # Dead-end elimination; trim each sub-list to the top 1000
            for EndPoint in PartialStrategiesByEndPoint:
                StrategyList=PartialStrategiesByEndPoint[EndPoint]
                # print(f'EndPoint {EndPoint}: {len(StrategyList)} Strategies')
                # If this list is greater than 1000, sort and trim to the top 1000
                if len(StrategyList)>MaxStrategies:
                    StrategyList.sort(key=lambda Strategy:scoreStrategy(Strategy)["total"],reverse=True) # Reverse order = higher scores first
                    StrategyList=StrategyList[0:MaxStrategies]

                # Pass group of strategies back to queue after trimming (or not trimming)
                for Strategy in StrategyList:
                    StrategyQueue.append(Strategy)

            # Report some info to screen at the end of each loop
            if report_to_screen==True:
                if len(PartialStrategiesByEndPoint)>0:
                    # Verbose printout for debugging
                    # print(f'Total number of endpoints = {len(PartialStrategiesByEndPoint)}, ranging from {min(PartialStrategiesByEndPoint.keys())} to {max(PartialStrategiesByEndPoint.keys())}')
                    # print(f'List of {len(NextQueue)} partial strategies trimmed to {len(StrategyQueue)}')
                    # User printout
                    report_trimmedstrategies=len(StrategyQueue)
                    print(f'End of loop: {report_finalstrats} complete strategies for {ProteinName}\n{report_trimmedstrategies} remaining {loopcount+1}-segment strategies')
                else:
                    print(f'No partial strategies left - this is the final loop.')

        # Finished with strategy-building loop
        if report_to_screen==True:
            gettime(f'--------\nDone with protein {ProteinName}...found {len(FinalStrategyList)} total strategies')

    # CLEAN DATA FOR FILE OUTPUT
    # Reset max strategies for file output
    if unrestrained_mode==True:
        MaxStrategies=OldMaxStrategies

    # Sort and trim final strategy list
    if StrategiesArePossible==True:
        FinalStrategyList.sort(key=lambda Strategy:scoreStrategy(Strategy)['total'],reverse=True)
        if len(FinalStrategyList)>MaxStrategies:
            FinalStrategyList = FinalStrategyList[0:MaxStrategies]
        if report_to_screen==True:
            gettime(f'Sorted final output list to top {MaxStrategies}')

        # Get the longest strategy in the list
        for Strategy in FinalStrategyList:
            StrategyLength=len(Strategy)-1 # Number of endpoints, minus the start 0
            if StrategyLength>MaxWidthSoFar:
                MaxWidthSoFar=StrategyLength
        MaxWidthDict[ProteinName]=MaxWidthSoFar

    print(f'Now writing to file')

    # OUTPUT TOTAL SCORES CSV FILE
    # Will be converted to .xlsx later
    OutputFilepath=f'{OutputFolder}/Aligator Analysis for {ProteinName}.csv'
    with open(OutputFilepath,'w') as f:
        # Write file header w/column names
        f.write('TOTAL SCORE,Thioester Score,Solubility Score,Segment Length Score,Thiol Penalty,#Ligations Penalty,Segments (from N- to C-terminus)...,\n')
        # If strategies were not possible, write N/A in relevant columns
        if StrategiesArePossible==False:
            f.write('n/a,n/a,n/a,n/a,n/a,n/a,NO STRATEGIES')
        # Otherwise, loop through final strategies and write to file
        else:
            for Strategy in FinalStrategyList:
                Scores=scoreStrategy(Strategy)
                f.write(f'{Scores["total"]},{Scores["thioester"]},{Scores["solubility"]},{Scores["length"]},{Scores["thiol"]},{Scores["ligations"]},')
                # Write AA sequence of each segment
                for i in range(0,len(Strategy)-1):
                    LeftIndex=Strategy[i]
                    RightIndex=Strategy[i+1]
                    f.write(ProteinSeq[LeftIndex:RightIndex]+',')
                f.write('\n')

    # Record for later
    LigFileDict[ProteinName]=OutputFilepath

    # OUTPUT ALL STRATEGIES TEXT FILE
    # Simpler output containing only the Total score and list of segments; ideal for passing to BracketMaker (github.com/Kay-Lab/BracketMaker)
    if output_all_strategies_text==True and StrategiesArePossible==True:
        # Name the folder and create the new directory (only first time)
        AllStrategiesFolder="Ligation Strategies Text Files"
        os.makedirs(f'{OutputFolder}/{AllStrategiesFolder}',exist_ok=True)
        with open(f'{OutputFolder}/{AllStrategiesFolder}/{ProteinName} All Strategies.txt','w') as f:
            # Write file header w/column names
            f.write("Strategy Score\tSegments\n")
            for Strategy in FinalStrategyList:
                Scores=scoreStrategy(Strategy)
                f.write(f'{Scores["total"]}\t')
                for i in range(0,len(Strategy)-1):
                    LeftIndex=Strategy[i]
                    RightIndex=Strategy[i+1]
                    SegmentKey=(LeftIndex,RightIndex)
                    SegmentSequence = SegmentScoreDict[SegmentKey]['seq']
                    f.write(f'{SegmentSequence}\t')
                f.write('\n')


    gettime('end')
    print("**************")
    print("")
    # END PROCESSING THIS FILE
# END LOOP THROUGH FILES

#Saves run time of Aligator.
RunTime = round((time.time() - start_time), 2)

#Writes run time to the run info file and closes the file.
RunInfoFile.write("Aligator took "+str(RunTime)+" seconds to run.")
RunInfoFile.close()

# MERGE AND FORMAT OUTPUT EXCEL DOCUMENTS
# Merge output .csv files into .xlsx documents
if merge_output_csv==True:
    print("Merging output CSV files to Excel format")

    # SEGMENT LISTS
    # Create Excel file for viable segment lists
    ExcelFileSeg=openpyxl.Workbook()

    # From recorded list of relevant files, create a new Excel sheet for each and transfer all of the data
    for SheetName in SegFileDict:
        ExcelFileSeg.create_sheet(index=-1,title=SheetName)
        sheet=ExcelFileSeg[SheetName]
        filepath=SegFileDict[SheetName]
        with open(filepath,'r') as f:
            reader=csv.reader(f, delimiter=',')
            for row in reader:
                sheet.append(row)

        # Apply formatting to header row
        for cell in ("A1","B1","C1","D1","E1","F1"):
            sheet[cell].alignment = center
            sheet[cell].font = Font(size = 12, bold = True)
        sheet["C1"].fill=aquaFill
        sheet["E1"].fill=redFill

        # Delete the csv file
        os.remove(filepath)

    # Remove initial blank sheet
    ExcelFileSeg.remove(ExcelFileSeg["Sheet"])

    # Save output file
    ExcelFileSeg.save(f"{OutputFolder}/Viable Segment Lists for {folder}.xlsx")


    # ALIGATOR STRATEGIES
    # Create Excel file for ligation strategy scores
    ExcelFileLig=openpyxl.Workbook()

    # From recorded list of relevant files, create a new Excel sheet for each and transfer all of the data
    for SheetName in LigFileDict:
        ExcelFileLig.create_sheet(index=-1,title=SheetName)
        sheet=ExcelFileLig[SheetName]
        filepath=LigFileDict[SheetName]
        with open(filepath,'r') as f:
            reader=csv.reader(f, delimiter=',')
            for row in reader:
                sheet.append(row)

        # Apply formatting to header row
        for cell in ("A1","B1","C1","D1","E1","F1","G1"):
            sheet[cell].alignment = center
            sheet[cell].font = Font(size = 12, bold = True)
        sheet["A1"].fill=greenFill
        sheet["B1"].fill=redFill
        sheet["C1"].fill=redFill
        sheet["D1"].fill=redFill
        sheet["E1"].fill=redFill
        sheet["F1"].fill=redFill
        sheet["G1"].fill=aquaFill

        # Merge "Segments N to C" header to span entire row, from width recorded during processing
        if MaxWidthDict[SheetName]>1:
            extracells=MaxWidthDict[SheetName]-1
            sheet.merge_cells(start_row=1, start_column=7, end_row=1, end_column=7+extracells)

        # Delete the csv file
        os.remove(filepath)

    # Remove initial blank sheet
    ExcelFileLig.remove(ExcelFileLig["Sheet"])

    # Save output file
    ExcelFileLig.save(f"{OutputFolder}/Aligator Analysis for {folder}.xlsx")

#Prints conclusion to user and lists full time it took to run Aligator.
print ("Aligator complete! Your data files are in the "+timestamp+" folder.")
print ("")
print ("Aligator took %s seconds to run." % RunTime)
